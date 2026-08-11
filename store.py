"""
store.py
--------
Reads the input Excel (universities/courses to research) and writes the
final professionally-formatted output workbook with three sheets:

    FINAL_RESULTS  - one row per (university, course, admission year)
    EVIDENCE       - every evidence snippet collected, with source metadata
    REVIEW_QUEUE   - only CROSS-CHECK REQUIRED / NOT FOUND rows

Uses openpyxl only (no pandas dependency) to keep the project lightweight.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule

import config

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


# --------------------------------------------------------------------------
# INPUT READING
# --------------------------------------------------------------------------
def read_input_rows(path: Path) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(
            f"Input file not found: {path}\n"
            f"Create it from the template described in README.md."
        )

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    def col(*names):
        for name in names:
            if name in header:
                return header.index(name)
        return None

    idx_univ = col("university name", "university", "colleges", "college")
    idx_site = col("official website", "official site", "website")
    idx_course = col("course", "course/programme", "programme")
    idx_year = col("admission year", "year")

    if idx_univ is None:
        raise ValueError("Input file must have a 'University Name' column.")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[idx_univ]:
            continue
        university = str(r[idx_univ]).strip()
        website = str(r[idx_site]).strip() if idx_site is not None and r[idx_site] else ""
        course = str(r[idx_course]).strip() if idx_course is not None and r[idx_course] else ""
        year_raw = r[idx_year] if idx_year is not None else None
        try:
            year = int(year_raw) if year_raw else 2026
        except (TypeError, ValueError):
            year = 2026

        if not course:
            continue  # course is required to scope the search

        rows.append({
            "university": university,
            "website": website,
            "course": course,
            "admission_year": year,
        })

    return rows


def read_existing_results(path: Path) -> list[dict]:
    """For resume support: load previously-written FINAL_RESULTS rows so a
    re-run doesn't lose earlier work. Returns [] if no output file yet."""
    if not path.exists():
        return []
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return []
    if "FINAL_RESULTS" not in wb.sheetnames:
        return []
    ws = wb["FINAL_RESULTS"]
    header = [c.value for c in ws[1]]
    key_map = {
        "University": "university", "Official Website": "official_website",
        "Course": "course", "Admission Year": "admission_year",
        "Admission Route": "admission_route", "Entrance Examination": "entrance_exam",
        "Exam Authority": "exam_authority", "Eligibility": "eligibility",
        "Selection Criteria": "selection_criteria",
        "Verification Status": "verification_status", "Confidence": "confidence",
        "Official Source": "source_title", "Source URL": "source_url",
        "Evidence": "evidence", "PDF Page": "pdf_page", "Checked On": "checked_date",
        "Notes": "notes",
    }
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for col_name, value in zip(header, row):
            key = key_map.get(col_name)
            if key:
                d[key] = value
        if d.get("university"):
            results.append(d)
    return results


# --------------------------------------------------------------------------
# OUTPUT WRITING
# --------------------------------------------------------------------------
FINAL_RESULTS_COLUMNS = [
    ("University", "university", 26),
    ("Official Website", "official_website", 30),
    ("Course", "course", 12),
    ("Admission Year", "admission_year", 14),
    ("Admission Route", "admission_route", 26),
    ("Entrance Examination", "entrance_exam", 20),
    ("Exam Authority", "exam_authority", 26),
    ("Eligibility", "eligibility", 24),
    ("Selection Criteria", "selection_criteria", 24),
    ("Verification Status", "verification_status", 18),
    ("Confidence", "confidence", 12),
    ("Official Source", "source_title", 26),
    ("Source URL", "source_url", 34),
    ("Evidence", "evidence", 46),
    ("PDF Page", "pdf_page", 10),
    ("Checked On", "checked_date", 14),
    ("Notes", "notes", 34),
]

EVIDENCE_COLUMNS = [
    ("University", "university", 26),
    ("Course", "course", 12),
    ("Admission Year", "admission_year", 14),
    ("Evidence", "text", 46),
    ("Source Title", "source_title", 26),
    ("Source URL", "source_url", 34),
    ("Source Type", "source_type", 16),
    ("PDF Page", "pdf_page", 10),
    ("Source Date", "source_date", 14),
    ("Official Source?", "is_official", 16),
    ("Relevance", "relevance", 12),
]

REVIEW_QUEUE_STATUSES = {config.STATUS_CROSS_CHECK, config.STATUS_NOT_FOUND}


def _write_sheet(ws, columns, rows, is_dataclass_list=False):
    headers = [c[0] for c in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in rows:
        values = []
        for _, key, _ in columns:
            val = getattr(row, key, None) if is_dataclass_list else row.get(key)
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            elif isinstance(val, list):
                val = ", ".join(val)
            values.append(val)
        ws.append(values)

    for i, (_, _, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP

    # Hyperlink the URL column(s)
    for i, (name, key, _) in enumerate(columns, start=1):
        if "url" in key.lower():
            for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
                for cell in row:
                    if cell.value:
                        cell.hyperlink = str(cell.value)
                        cell.font = Font(color="0563C1", underline="single")

    last_col_letter = get_column_letter(len(columns))
    last_row = max(ws.max_row, 1)
    if last_row >= 2:
        table = Table(displayName=ws.title.replace(" ", "_"), ref=f"A1:{last_col_letter}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(table)


def _apply_status_conditional_formatting(ws, status_col_letter: str, last_row: int):
    if last_row < 2:
        return
    rng = f"{status_col_letter}2:{status_col_letter}{last_row}"
    for status, color in config.STATUS_COLORS.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill)
        )


def write_workbook(path: Path, results: list[dict], evidence: list) -> None:
    wb = Workbook()

    # --- FINAL_RESULTS ---
    ws1 = wb.active
    ws1.title = "FINAL_RESULTS"
    _write_sheet(ws1, FINAL_RESULTS_COLUMNS, results)
    status_col_idx = [k for _, k, _ in FINAL_RESULTS_COLUMNS].index("verification_status") + 1
    status_col_letter = get_column_letter(status_col_idx)
    _apply_status_conditional_formatting(ws1, status_col_letter, ws1.max_row)
    ws1.auto_filter.ref = ws1.dimensions

    # --- EVIDENCE ---
    ws2 = wb.create_sheet("EVIDENCE")
    _write_sheet(ws2, EVIDENCE_COLUMNS, evidence, is_dataclass_list=True)
    ws2.auto_filter.ref = ws2.dimensions

    # --- REVIEW_QUEUE ---
    ws3 = wb.create_sheet("REVIEW_QUEUE")
    review_rows = [r for r in results if r.get("verification_status") in REVIEW_QUEUE_STATUSES]
    _write_sheet(ws3, FINAL_RESULTS_COLUMNS, review_rows)
    if review_rows:
        _apply_status_conditional_formatting(ws3, status_col_letter, ws3.max_row)
        ws3.auto_filter.ref = ws3.dimensions

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
