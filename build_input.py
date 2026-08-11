"""
build_input.py
---------------
One-off helper: converts a simple "College Name | Official Site" list (like
the collage-list.xlsx you already have) into the project's required input
format, expanding each college into one row per course.

Usage:
    python build_input.py path/to/collage-list.xlsx --courses "B.Tech,MBA,BBA" --year 2026

Not part of the core pipeline (scraper.py / checker.py / store.py) - this is
just a convenience so you don't have to retype 65 rows x 3 courses by hand.
"""

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook


def build(source_path: Path, dest_path: Path, courses: list[str], year: int) -> int:
    src = load_workbook(source_path, data_only=True)
    ws = src.active

    wb = Workbook()
    out = wb.active
    out.title = "Sheet1"
    out.append(["University Name", "Official Website", "Course", "Admission Year"])

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        university = str(row[0]).strip()
        website = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        for course in courses:
            out.append([university, website, course, year])
            count += 1

    dest_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest_path)
    return count


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path, help="Path to your College/Site list (.xlsx)")
    parser.add_argument("--dest", type=Path, default=Path("input/universities.xlsx"))
    parser.add_argument("--courses", type=str, default="B.Tech,MBA,BBA",
                         help="Comma-separated list of courses to check per college")
    parser.add_argument("--year", type=int, default=2026)
    args = parser.parse_args()

    course_list = [c.strip() for c in args.courses.split(",") if c.strip()]
    n = build(args.source, args.dest, course_list, args.year)
    print(f"Wrote {n} rows to {args.dest}")
